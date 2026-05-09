# base list screen class
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.screen import MDScreen
from kivymd.uix.label import MDLabel
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.divider import MDDivider
from kivymd.uix.button import MDIconButton,MDButton,MDButtonText
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.widget import MDWidget
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.dialog import (
    MDDialog,
    MDDialogIcon,
    MDDialogHeadlineText,
    MDDialogSupportingText,
    MDDialogButtonContainer,
)


from kivy.storage.jsonstore import JsonStore
from kivy.utils import platform
from kivy.metrics import dp

import os
import openpyxl

class TableRow(MDBoxLayout):
    orientation= 'horizontal'
    def __init__(self, parent, cols: list, data:dict, col_id:str, **kwargs):
        super().__init__(**kwargs)
        
        self.cols = cols
        self._id = col_id
        self._parent = parent

        self.size_hint_y = None
        self.height = dp(48)

        self.size_hint_x = None
        self.width = dp(120) * len(cols)

        layout = MDBoxLayout()
        layout.bind(minimum_width=layout.setter("width"))
        
        if self._id != -1:
            options = MDIconButton(icon="dots-vertical")
                
            menu_items = [
                {"text": "Edit","on_release": lambda: self.menu_callback("Edit")},
                {"text": "Delete","on_release": lambda: self.menu_callback("Delete")},
            ]
            
            self.menu = MDDropdownMenu(
                caller=options,
                items=menu_items,
            )
            
            options.bind(on_release=lambda x: self.menu.open())
        
        else:
            options = MDBoxLayout(width=40.0,size_hint=(None,None))
        
        layout.add_widget(options)
        
        for i in self.cols:
            l = MDLabel(
                text=str(data.get(i, "")),
                size_hint_x=None,
                width=dp(120),
                shorten=True,
                padding=[10,0]
                )
            layout.add_widget(l)
            
            layout.add_widget(MDDivider(orientation="vertical"))

        self.add_widget(layout)
    
    def menu_callback(self, action):
        self.menu.dismiss()
        if action=="Delete":
            self._parent.remove_item(self._id,self)

class ListScreen(MDScreen):
    store:JsonStore
    path:str
    df:dict
    cols = ['FN', 'LN', 'val', 'exp', 'iss', 'UID', 'NO', 'CID', 'ID', 'DOB', 'NAT', 'TM', 'LT', 'LTN', 'LT2', 'KEY']
    def __init__(self, path, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        if platform == "android":
            self.path = os.path.join(path, "data.json")
        else:
            self.path = os.path.join(path, "test_app", "data.json")
        
        self.store = JsonStore(self.path)
        
        self.get_data()
        self.cleanup_numbering()
        
        self.manager_open = False
        self.file_manager = MDFileManager(
            exit_manager=self.exit_manager, select_path=self.select_path, selector="folder"
        )
        
        #layout
        self.box = MDBoxLayout(orientation= 'vertical', size_hint=(1,1))
        
        self.box.add_widget(
            MDBoxLayout(
                MDIconButton(icon="trash-can-outline",on_release=lambda _: self.delete_dialog()),
                MDWidget(),
                MDIconButton(icon="export-variant",on_release=lambda _: self.file_manager_open()),
                orientation= 'horizontal',
                size_hint_y=None,
                height=dp(48),
                padding=dp(5)
            )
        )
        
        self.rv = MDScrollView(do_scroll_x=True,do_scroll_y=True,size_hint_x=1)

        self.layout = MDBoxLayout(
            orientation="vertical",
            size_hint=(None,None)
        )
        self.layout.bind(
            minimum_height=self.layout.setter("height"),
            minimum_width=self.layout.setter("width")
        )
        
        self.setup_columns()
        
        for i,v in reversed(self.df.items()):
            self.layout.add_widget(TableRow(self, self.cols, v, i))
        
        self.rv.add_widget(self.layout)
    
        self.box.add_widget(self.rv)
        
        self.add_widget(self.box)
        
        self._print_data()
    
    def setup_columns(self):
        self.layout.add_widget(TableRow(self, self.cols, {c:c for c in self.cols},-1))
    
    def get_data(self):
        if self.store.exists("df"):
            self.df = self.store.get("df").get("values")
            if not isinstance(self.df,dict):
                self.df = {}
            elif len(list(self.df.values())) < 1:
                self.df = {}
            elif set(list(self.df.values())[0].keys()) != set(self.cols):
                self.df = {}
        else:
            self.df = {}
        
        self.save_data()
    
    def save_data(self):
        self.store.put(key='df', values=self.df)
    
    def cleanup_numbering(self):
        val = self.df.values()
        self.df = {}
        for i,v in enumerate(val):
            self.df[str(i)] = v
    
    
    def file_manager_open(self):
        self.file_manager.show(
            os.path.expanduser("~"))  # output manager to the screen
        self.manager_open = True

    def exit_manager(self, *args):
        '''Called when the user reaches the root of the directory tree.'''

        self.manager_open = False
        self.file_manager.close()

    def select_path(self, path: str):
        '''
        It will be called when you click on the file name
        or the catalog selection button.

        :param path: path to the selected directory or file;
        '''

        self.exit_manager()
        
        self.cleanup_numbering()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Judopass_data"
        
        for i,v in enumerate(self.cols):
            ws.cell(1,i+1,v)
        
        for k in self.df:
            for i,v in enumerate(self.cols):
                value = str(self.df[k][v])
                
                if value.isdecimal():
                    value = int(value)
                
                ws.cell(int(k)+2,i+1,value)
            
        
        wb.save(os.path.join(path,"judopass_export.xlsx"))


    def delete_dialog(self):
        self._dialog = MDDialog(
            MDDialogIcon(icon="trash-can-outline"),
            MDDialogHeadlineText(text="Are you sure?"),
            MDDialogSupportingText(text="This Action will delete all avalabile data"),
            MDDialogButtonContainer(
                MDWidget(),
                MDButton(MDButtonText(text="Cancel"),style="text",on_release=lambda _:self._dialog.dismiss()),
                MDButton(MDButtonText(text="Continue"),style="text",on_release=lambda _:self.delete_data())
            )
        )
        
        self._dialog.open()
    
    def delete_data(self):
        if self._dialog:
            self._dialog.dismiss()
        
        self.layout.clear_widgets()
        self.df = {}
        self.setup_columns()
        #self.save_data()


    def set_data(self,data):
        keys = self.df.keys()
        if len(keys) < 1:
            id = 1
        else:
            id = str(int(max(keys))+1)
        
        self.df[id] = data
        
        self.layout.add_widget(TableRow(self,self.cols,data,id),-1)
        
        self.save_data()
        
        self._print_data()

        self.save_data()
    
    def _print_data(self):
        import json
        print(json.dumps(self.df,indent=4))
    
    def remove_item(self,id,item):
        self.df.pop(id)
        self.save_data()
        self.layout.remove_widget(item)
        